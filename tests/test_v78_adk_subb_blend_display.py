import importlib.util
import io
import inspect
from pathlib import Path

import openpyxl
import numpy as np
import pandas as pd


def load_v78_module():
    root = Path(__file__).resolve().parents[1]
    path = root / "mnt_bot V 7.8 plus.py"
    spec = importlib.util.spec_from_file_location("mnt_bot_v78_plus", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_v78_dk_full_pool_frame_clamps_to_zz1000_publication_date():
    module = load_v78_module()
    dates = pd.to_datetime(["2014-10-16", "2014-10-17", "2014-10-20"])
    dk_dfs = {
        col: pd.DataFrame({col: [1.0, 2.0, 3.0]}, index=dates)
        for col in module.CN_DK_COLS
    }

    out = module._build_cn_dk_close_frame(dk_dfs)

    assert out.index.min() == pd.Timestamp("2014-10-17")


def test_v78_adk_status_rows_include_v77_and_new_adk_legs():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame(
        {"top_pair": ["ZZ1000/SZ50"], "direction": [1], "weight": [0.8]},
        index=dates,
    )
    v77.attrs["signals_df"] = pd.DataFrame({"ZZ1000/SZ50": [55.0]}, index=dates)
    new = pd.DataFrame(
        {
            "top_pair": ["ZZ500/CYB"],
            "direction": [-1],
            "weight": [0.0],
            "v78_score_overheat_score": [92.0],
            "v78_score_overheat_scale": [0.0],
            "v78_score_overheat_on": [True],
        },
        index=dates,
    )
    new.attrs["signals_df"] = pd.DataFrame({"ZZ500/CYB": [92.0]}, index=dates)
    blended = pd.DataFrame({"v78_adk_v77_holding": ["ZZ1000/SZ50_1"]}, index=dates)
    blended.attrs["v78_adk_v77"] = v77
    blended.attrs["v78_adk_new"] = new

    rows = module._v78_adk_leg_status_rows(blended, 0)

    assert [row["leg"] for row in rows] == ["V7.7 ADK", module.V78_ADK_NEW_LABEL]
    assert rows[0]["pair"] == "ZZ1000/SZ50"
    assert rows[0]["score"] == 55.0
    assert rows[1]["pair"] == "ZZ500/CYB"
    assert rows[1]["score"] == 92.0
    assert rows[1]["score_hot_on"] is True
    assert rows[1]["leg_contribution"] == 0.0


def test_v78_adk_blend_exposes_final_leg_components_and_signal_flags():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-11", "2026-06-12"])
    v77 = pd.DataFrame(
        {
            "return": [0.0, 0.0],
            "holding": ["SZ50/ZZ1000_1", "SZ50/ZZ1000_1"],
            "top_pair": ["SZ50/ZZ1000", "SZ50/ZZ1000"],
            "long_leg": ["SZ50", "SZ50"],
            "short_leg": ["ZZ1000", "ZZ1000"],
            "direction": [1, 1],
            "weight": [1.0, 1.0],
            "is_signal": [False, False],
        },
        index=dates,
    )
    new = pd.DataFrame(
        {
            "return": [0.0, 0.0],
            "holding": ["HS300/CYB_1", "HS300/CYB_1"],
            "top_pair": ["HS300/CYB", "HS300/CYB"],
            "long_leg": ["HS300", "HS300"],
            "short_leg": ["CYB", "CYB"],
            "direction": [1, 1],
            "weight": [0.2, 0.6],
            "is_signal": [False, False],
        },
        index=dates,
    )

    out = module.blend_v78_adk_results(v77, new)

    assert np.allclose(out["final_exposure"], [0.6, 0.8])
    assert np.allclose(out["weight"], out["final_exposure"])
    assert out.loc[dates[0], "holding"] == "V7.7ADK:SZ50/ZZ1000_1|NewADK:HS300/CYB_1"
    assert out.loc[dates[0], "final_long_short_legs"]["v77"]["long_leg"] == "SZ50"
    assert out.loc[dates[0], "final_long_short_legs"]["new"]["short_leg"] == "CYB"
    assert out.loc[dates[0], "adk_net_asset_exposure"] == {
        "CYB": -0.1,
        "HS300": 0.1,
        "SZ50": 0.5,
        "ZZ1000": -0.5,
    }
    assert "V7.7ADK" in module._dk_pos_str(out.loc[dates[0], "holding"])
    assert "NewADK" in module._dk_pos_str(out.loc[dates[0], "holding"])
    assert bool(out.loc[dates[1], "is_signal"]) is True


def test_v78_subb_four_leg_rows_expand_nested_v77_components():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame(
        {
            "official_w_QQQ": [0.60],
            "official_contrib_w_QQQ": [0.30],
            "ema_w_QQQ": [0.40],
            "ema_contrib_w_QQQ": [0.20],
            "w_QQQ": [0.50],
        },
        index=dates,
    )
    bias = pd.DataFrame({"w_QQQ": [0.20]}, index=dates)
    logvol = pd.DataFrame({"w_QQQ": [0.10]}, index=dates)
    blended = pd.DataFrame({"w_QQQ": [0.30]}, index=dates)
    blended.attrs["v78_subb_v77"] = v77
    blended.attrs["v78_subb_bias"] = bias
    blended.attrs["v78_subb_logvol"] = logvol

    rows = module._v78_subb_four_leg_weight_rows(blended, 0)

    assert rows == [
        {
            "asset": "QQQ",
            "live_name": "QQQM",
            "official_contrib": 0.15,
            "ema_contrib": 0.10,
            "bias_contrib": 0.05,
            "logvol_contrib": 0.025,
            "final_weight": 0.325,
        }
    ]


def test_v78_subb_blend_targets_use_leg_targets_not_current_holdings():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame(
        {
            "return": [0.0],
            "w_QQQ": [0.10],
            "actual_w_QQQ": [0.10],
            "target_w_QQQ": [0.50],
            "official_w_QQQ": [0.60],
            "official_contrib_w_QQQ": [0.30],
            "ema_w_QQQ": [0.40],
            "ema_contrib_w_QQQ": [0.20],
            "is_signal": [True],
            "rebalanced": [False],
        },
        index=dates,
    )
    bias = pd.DataFrame(
        {
            "return": [0.0],
            "w_QQQ": [0.00],
            "actual_w_QQQ": [0.00],
            "target_w_QQQ": [0.40],
            "is_signal": [True],
            "rebalanced": [True],
        },
        index=dates,
    )
    logvol = pd.DataFrame(
        {
            "return": [0.0],
            "w_QQQ": [0.00],
            "actual_w_QQQ": [0.00],
            "target_w_QQQ": [0.20],
            "is_signal": [True],
            "rebalanced": [True],
        },
        index=dates,
    )

    blended = module.blend_v78_subb_results(v77, bias, logvol)
    rows = module._v78_subb_four_leg_weight_rows(blended, dates[0])
    signal_weights = module._subb_signal_display_source_weights(blended, dates[0], ["w_QQQ"])

    assert round(float(blended.loc[dates[0], "actual_w_QQQ"]), 10) == 0.05
    assert round(float(blended.loc[dates[0], "target_w_QQQ"]), 10) == 0.40
    assert bool(blended.loc[dates[0], "rebalanced"]) is True
    assert round(signal_weights["QQQ"], 10) == 0.40
    assert round(rows[0]["official_contrib"], 10) == 0.15
    assert round(rows[0]["ema_contrib"], 10) == 0.10
    assert round(rows[0]["bias_contrib"], 10) == 0.10
    assert round(rows[0]["logvol_contrib"], 10) == 0.05
    assert round(rows[0]["final_weight"], 10) == 0.40


def test_v78_subb_blend_syncs_model_rebalanced_with_any_leg_rebalance():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame(
        {"return": [0.0], "w_QQQ": [0.10], "rebalanced": [False], "model_rebalanced": [False]},
        index=dates,
    )
    bias = pd.DataFrame({"return": [0.0], "w_QQQ": [0.20], "rebalanced": [True]}, index=dates)
    logvol = pd.DataFrame({"return": [0.0], "w_QQQ": [0.30], "rebalanced": [False]}, index=dates)

    blended = module.blend_v78_subb_results(v77, bias, logvol)

    assert bool(blended.loc[dates[0], "rebalanced"]) is True
    assert bool(blended.loc[dates[0], "model_rebalanced"]) is True
    assert bool(blended.loc[dates[0], "effective_rebalanced"]) is True


def test_v78_subb_blend_resets_legacy_execution_cost_fields():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame(
        {
            "return": [0.01],
            "return_before_subb_execution_cost": [0.99],
            "subb_execution_cost": [0.88],
            "w_QQQ": [0.10],
        },
        index=dates,
    )
    bias = pd.DataFrame({"return": [0.05], "w_QQQ": [0.20]}, index=dates)
    logvol = pd.DataFrame({"return": [0.03], "w_QQQ": [0.30]}, index=dates)

    blended = module.blend_v78_subb_results(v77, bias, logvol)

    assert abs(float(blended.loc[dates[0], "return_before_subb_execution_cost"]) - float(blended.loc[dates[0], "return"])) < 1e-12
    assert float(blended.loc[dates[0], "subb_execution_cost"]) == 0.0
    assert abs(float(blended.loc[dates[0], "v78_subb_component_net_return"]) - float(blended.loc[dates[0], "return"])) < 1e-12
    assert "component-net blend" in blended.loc[dates[0], "cost_basis_note"]
    assert "display-only" in blended.loc[dates[0], "cost_basis_note"]


def test_v78_suba_and_adk_blends_reset_legacy_execution_cost_fields():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77_suba = pd.DataFrame(
        {
            "return": [0.01],
            "holding": ["AAA"],
            "weight": [1.0],
            "trade_cost": [0.88],
            "turnover": [0.77],
            "effective_turnover": [0.66],
        },
        index=dates,
    )
    new_suba = pd.DataFrame({"return": [0.03], "holding": ["BBB"], "weight": [0.5]}, index=dates)
    suba = module.blend_v78_suba_results(v77_suba, new_suba)

    assert abs(float(suba.loc[dates[0], "v78_suba_component_net_return"]) - float(suba.loc[dates[0], "return"])) < 1e-12
    assert abs(float(suba.loc[dates[0], "return_before_suba_execution_cost"]) - float(suba.loc[dates[0], "return"])) < 1e-12
    assert float(suba.loc[dates[0], "trade_cost"]) == 0.0
    assert np.isnan(float(suba.loc[dates[0], "turnover"]))
    assert np.isnan(float(suba.loc[dates[0], "effective_turnover"]))

    v77_adk = pd.DataFrame(
        {
            "return": [0.01],
            "holding": ["SZ50/ZZ1000_1"],
            "top_pair": ["SZ50/ZZ1000"],
            "long_leg": ["SZ50"],
            "short_leg": ["ZZ1000"],
            "weight": [1.0],
            "return_before_dk_execution_cost": [0.99],
            "dk_execution_cost": [0.88],
            "dk_execution_turnover": [0.77],
        },
        index=dates,
    )
    new_adk = pd.DataFrame(
        {
            "return": [0.03],
            "holding": ["HS300/CYB_1"],
            "top_pair": ["HS300/CYB"],
            "long_leg": ["HS300"],
            "short_leg": ["CYB"],
            "weight": [0.5],
        },
        index=dates,
    )
    adk = module.blend_v78_adk_results(v77_adk, new_adk)

    assert abs(float(adk.loc[dates[0], "v78_adk_component_net_return"]) - float(adk.loc[dates[0], "return"])) < 1e-12
    assert abs(float(adk.loc[dates[0], "return_before_dk_execution_cost"]) - float(adk.loc[dates[0], "return"])) < 1e-12
    assert float(adk.loc[dates[0], "dk_execution_cost"]) == 0.0
    assert np.isnan(float(adk.loc[dates[0], "dk_execution_turnover"]))


def test_subb_regular_rebalance_records_skip_volreg_cash_model_changes():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-11", "2026-06-12"])
    result = pd.DataFrame(
        {
            "rebalanced": [False, True],
            "model_rebalanced": [False, True],
            "volreg_cash": [False, True],
            "target_w_QQQ": [0.50, 0.00],
            "target_w_GLD": [0.50, 1.00],
            "actual_w_QQQ": [0.50, 0.00],
            "actual_w_GLD": [0.50, 0.00],
            "actual_w_BIL": [0.00, 1.00],
        },
        index=dates,
    )

    records = module.extract_us_rot_rebalances(result, since_date=pd.Timestamp("2026-06-12"))

    assert records == []


def test_subb_regular_rebalance_records_skip_volreg_transition_model_changes():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-11", "2026-06-12"])
    result = pd.DataFrame(
        {
            "rebalanced": [False, True],
            "model_rebalanced": [False, True],
            "volreg_cash": [True, False],
            "volreg_transition": [False, True],
            "volreg_action": ["", "exit_cash"],
            "target_w_QQQ": [0.50, 0.60],
            "target_w_GLD": [0.50, 0.40],
            "actual_w_QQQ": [0.00, 0.60],
            "actual_w_GLD": [0.00, 0.40],
            "actual_w_BIL": [1.00, 0.00],
        },
        index=dates,
    )

    records = module.extract_us_rot_rebalances(result, since_date=pd.Timestamp("2026-06-12"))

    assert records == []


def test_v78_logvol_high_vol_scales_target_weights_not_return_discount(monkeypatch):
    module = load_v78_module()
    dates = pd.bdate_range("2024-01-02", periods=390)
    close = pd.DataFrame(index=dates)
    for code in module.US_ROT_POOL:
        close[code] = 100.0
    close["QQQ"] = 100.0 * np.exp(np.where(np.arange(len(dates)) % 2 == 0, 0.06, -0.06).cumsum())
    close["BIL"] = 100.0
    close["SPY"] = 100.0
    dummy_scores = pd.DataFrame(1.0, index=dates, columns=module.US_ROT_POOL)

    monkeypatch.setattr(module, "_v78_score_log_weighted", lambda close_df: dummy_scores)
    monkeypatch.setattr(module, "_v78_target_from_scores", lambda *args, **kwargs: {"QQQ": 1.0})
    monkeypatch.setattr(module, "_v78_spy_volume_gate", lambda index: (pd.Series(False, index=index), "test volume"))
    monkeypatch.setattr(module, "_us_signal_days", lambda close_df, start_idx: set(range(start_idx, len(close_df))))
    monkeypatch.setattr(module, "US_ROT_COMMISSION", 0.0)

    out = module.run_v78_subb_new_line(close, line="logvol")
    high_vol_signals = out[out["is_signal"] & out["logvol_high_vol_on"]]

    assert not high_vol_signals.empty
    row = high_vol_signals.iloc[0]
    assert float(row["logvol_high_vol_scale"]) == 0.75
    assert abs(float(row["target_w_QQQ"]) - 0.75) < 1e-12
    assert abs(float(row["target_w_BIL"]) - 0.25) < 1e-12


def test_v78_subb_volreg_is_applied_after_final_v78_blend():
    module = load_v78_module()
    source = inspect.getsource(module.CombinedStrategyBase._run_strategies)

    blend_pos = source.index("us_rot_result = blend_v78_subb_results")
    volreg_pos = source.index("us_rot_result = apply_vol_regime_overlay")

    assert blend_pos < volreg_pos


def test_subb_volreg_preserves_v78_return_when_not_cash():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    result = pd.DataFrame(
        {
            "return": [0.0300],
            "return_before_subb_execution_cost": [0.0100],
            "v78_subb_v77_return": [0.0100],
            "v78_subb_bias_return": [0.0500],
            "v78_subb_logvol_return": [0.0300],
            "actual_w_QQQ": [1.0],
            "rebalanced": [False],
        },
        index=dates,
    )
    spy_close = pd.Series([100.0], index=dates)

    out = module.apply_vol_regime_overlay(result, spy_close)

    assert bool(out.loc[dates[0], "volreg_cash"]) is False
    assert abs(float(out.loc[dates[0], "return"]) - 0.0300) < 1e-12
    assert abs(float(out.loc[dates[0], "gross_return_before_volreg_cost"]) - 0.0300) < 1e-12


def test_v78_subb_hypothetical_weights_blend_all_four_legs():
    module = load_v78_module()

    weights = module._blend_v78_subb_weight_dicts(
        {"QQQ": 0.50},
        {"PDBC": 0.40},
        {"GLD": 0.20},
    )

    assert weights == {"GLD": 0.05, "PDBC": 0.10, "QQQ": 0.25}


def test_v78_adk_score_hot_rebuilds_execution_costs_after_scaling():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-11", "2026-06-12"])
    dk_result = pd.DataFrame(
        {
            "return": [0.0100, 0.0200],
            "return_before_dk_execution_cost": [0.0100, 0.0200],
            "top_pair": ["SZ50/HS300", "SZ50/HS300"],
            "direction": [1, 1],
            "weight": [1.0, 1.0],
        },
        index=dates,
    )
    dk_result.attrs["signals_df"] = pd.DataFrame({"SZ50/HS300": [100.0, 100.0]}, index=dates)
    dk_result.attrs["pair_data"] = {
        "SZ50/HS300": pd.DataFrame(
            {"position": [1, 1], "raw_ret": [0.0100, 0.0200]},
            index=dates,
        )
    }

    out = module.apply_v78_adk_score_overheat(dk_result, enter=80.0, exit=20.0, derisk_scale=0.5)

    assert "dk_execution_cost" in out.columns
    assert out["dk_execution_cost"].sum() > 0.0
    assert out.loc[dates[1], "weight"] == 0.5
    assert out.loc[dates[1], "return"] < 0.0200 * 0.5


def test_dk_pos_str_composite_holding_has_no_mojibake():
    module = load_v78_module()

    text = module._dk_pos_str("V7.7ADK:HS300/ZZ500_1|NewADK:SZ50/CYB_-1")

    assert "V7.7ADK" in text
    assert "NewADK" in text
    assert "閸" not in text
    assert "?" not in text


def test_v78_adk_rebalances_use_component_legs_not_composite_parser():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-11", "2026-06-12"])
    blended = pd.DataFrame(
        {
            "holding": [
                "V7.7ADK:SZ50/ZZ1000_1|NewADK:HS300/CYB_1",
                "V7.7ADK:HS300/ZZ1000_1|NewADK:HS300/CYB_1",
            ],
            "weight": [1.0, 1.0],
        },
        index=dates,
    )
    v77 = pd.DataFrame(
        {
            "holding": ["SZ50/ZZ1000_1", "HS300/ZZ1000_1"],
            "weight": [1.0, 1.0],
        },
        index=dates,
    )
    new = pd.DataFrame(
        {
            "holding": ["HS300/CYB_1", "HS300/CYB_1"],
            "weight": [1.0, 1.0],
        },
        index=dates,
    )
    blended.attrs["v78_adk_v77"] = v77
    blended.attrs["v78_adk_new"] = new

    legacy_records = module.extract_dk_rebalances(blended, strategy_name="V7.8 ADK")
    v78_records = module.extract_v78_adk_rebalances(blended)

    assert legacy_records == []
    assert len(v78_records) == 1
    assert "策略" in v78_records[0]
    v78_records[0]["\u7edb\u682b\u6690"] = v78_records[0]["策略"]
    assert "V7.7 ADK" in v78_records[0]["绛栫暐"]


def test_v78_adk_rebalances_since_date_and_sort_use_date_key():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-10", "2026-06-11", "2026-06-12"])
    blended = pd.DataFrame({"holding": ["x", "y", "z"], "weight": [1.0, 1.0, 1.0]}, index=dates)
    v77 = pd.DataFrame(
        {
            "holding": ["SZ50/ZZ1000_1", "HS300/ZZ1000_1", "ZZ500/ZZ1000_1"],
            "weight": [1.0, 1.0, 1.0],
        },
        index=dates,
    )
    new = pd.DataFrame(
        {
            "holding": ["HS300/CYB_1", "HS300/CYB_1", "SZ50/CYB_1"],
            "weight": [1.0, 1.0, 1.0],
        },
        index=dates,
    )
    blended.attrs["v78_adk_v77"] = v77
    blended.attrs["v78_adk_new"] = new

    records = module.extract_v78_adk_rebalances(blended, since_date=pd.Timestamp("2026-06-11"))

    assert records
    assert all(pd.Timestamp(record["日期"]) >= pd.Timestamp("2026-06-11") for record in records)
    assert [record["日期"] for record in records] == sorted(record["日期"] for record in records)
    assert all("閺" not in key for record in records for key in record)


def test_subb_rebalance_extract_since_date_uses_previous_row_as_baseline():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-05-01", "2026-06-12"])
    result = pd.DataFrame(
        {
            "rebalanced": [False, True],
            "target_w_QQQ": [0.25, 0.40],
            "target_w_GLD": [0.75, 0.60],
            "actual_w_QQQ": [0.25, 0.25],
            "actual_w_GLD": [0.75, 0.75],
        },
        index=dates,
    )

    records = module.extract_us_rot_rebalances(result, since_date=pd.Timestamp("2026-06-01"))

    assert len(records) == 1
    assert records[0]["日期"] == "2026-06-12"
    assert "QQQM 25.0%->40.0%" in records[0]["买入"]
    assert "GLDM 75.0%->60.0%" in records[0]["卖出"]


def test_subb_volreg_extract_since_date_uses_previous_row_as_baseline():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-05-01", "2026-06-12"])
    result = pd.DataFrame(
        {
            "volreg_transition": [False, True],
            "effective_w_QQQ": [0.50, 0.0],
            "effective_w_GLD": [0.50, 0.0],
            "effective_w_BIL": [0.0, 1.0],
            "w_QQQ": [0.50, 0.50],
            "w_GLD": [0.50, 0.50],
            "w_BIL": [0.0, 0.0],
        },
        index=dates,
    )

    records = module.extract_subb_volreg_rebalances(result, since_date=pd.Timestamp("2026-06-01"))

    assert len(records) == 1
    assert records[0]["日期"] == "2026-06-12"
    assert "QQQM 50.0%->0.0%" in records[0]["卖出"]
    assert "GLDM 50.0%->0.0%" in records[0]["卖出"]


def test_v78_subb_component_leg_tables_show_each_leg_without_legacy_hypothetical_block():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame(
        {
            "official_w_QQQ": [0.60],
            "official_contrib_w_QQQ": [0.30],
            "ema_w_QQQ": [0.40],
            "ema_contrib_w_QQQ": [0.20],
            "w_QQQ": [0.50],
        },
        index=dates,
    )
    bias = pd.DataFrame({"w_PDBC": [0.44]}, index=dates)
    logvol = pd.DataFrame({"w_GLDM": [0.30]}, index=dates)
    blended = pd.DataFrame({"w_QQQ": [0.25], "w_PDBC": [0.11], "w_GLDM": [0.075]}, index=dates)
    blended.attrs["v78_subb_v77"] = v77
    blended.attrs["v78_subb_bias"] = bias
    blended.attrs["v78_subb_logvol"] = logvol
    chunks = []

    module._write_v78_subb_component_leg_tables(chunks.append, blended, 0)
    text = "".join(chunks)

    assert "**EMA腿（25%）**" in text
    assert "**Bias腿（25%）**" in text
    assert "**LogVol腿（25%）**" in text
    assert "QQQM" in text
    assert "PDBC" in text
    assert "GLDM" in text
    assert "假设收盘信号" not in text


def test_v78_subb_leg_tables_accept_signal_date_index():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame(
        {
            "official_w_QQQ": [0.60],
            "official_contrib_w_QQQ": [0.30],
            "ema_w_QQQ": [0.40],
            "ema_contrib_w_QQQ": [0.20],
            "w_QQQ": [0.50],
        },
        index=dates,
    )
    bias = pd.DataFrame({"w_PDBC": [0.44]}, index=dates)
    logvol = pd.DataFrame({"w_GLDM": [0.30]}, index=dates)
    blended = pd.DataFrame({"w_QQQ": [0.25], "w_PDBC": [0.11], "w_GLDM": [0.075]}, index=dates)
    blended.attrs["v78_subb_v77"] = v77
    blended.attrs["v78_subb_bias"] = bias
    blended.attrs["v78_subb_logvol"] = logvol
    chunks = []

    module._write_v78_subb_component_leg_tables(chunks.append, blended, dates[0])
    module._write_v78_subb_blend_table(chunks.append, blended, dates[0])
    text = "".join(chunks)

    assert "QQQM" in text
    assert "PDBC" in text
    assert "GLDM" in text


def test_v78_subb_current_vs_hypothetical_table_compares_actual_and_today_target():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-07-02"])
    blended = pd.DataFrame(
        {
            "v78_subb_v77_return": [0.0],
            "actual_w_PDBC": [0.50],
            "actual_w_QQQ": [0.30],
            "actual_w_EMXC": [0.20],
            "target_w_PDBC": [0.447],
            "target_w_QQQ": [0.381],
            "target_w_EMXC": [0.193],
            "target_w_GLD": [0.010],
        },
        index=dates,
    )
    chunks = []

    module._write_v78_subb_current_vs_hypothetical_table(chunks.append, blended, 0)
    text = "".join(chunks)

    assert "**V7.8 Sub-B 当前持有 vs 假设今日调仓**" in text
    assert "| PDBC | 50.0% | **44.7%** | -5.3% |" in text
    assert "| QQQM | 30.0% | **38.1%** | +8.1% |" in text
    assert "| GLDM | 0.0% | **1.0%** | +1.0% |" in text
    assert "假设今日是Sub-B调仓日" in text


def test_v78_subb_param_tables_list_each_leg_separately():
    module = load_v78_module()
    chunks = []

    module._write_v78_subb_param_tables(chunks.append)
    text = "".join(chunks)

    assert "**全局执行口径**" in text
    assert "**官方腿参数**" in text
    assert "**EMA腿参数**" in text
    assert "**Bias腿参数**" in text
    assert "**LogVol腿参数**" in text
    assert "hl100 / 16%" in text
    assert "ewma6m_1vol" in text
    assert "price/MA160,260,390 = 3/2/1加权" in text
    assert "25% / 40日 / max1.5x" in text
    assert "log return 120/200/320 = 60%/30%/10%" in text
    assert "30% / 40日 / max1.25x" in text
    assert "SPY量/MA60≥1.5 -> QQQ/EMXC/EFA ×0.75" in text
    assert "component-net" in text
    assert "V7.7 official+EMA account-level blend is pre-netted before entering V7.8 as the 50% V7.7 component" in text
    assert "fail_closed" in text
    assert "local runs without Yahoo access may be more defensive" in text
    assert "rv≥50% -> QQQ/EMXC/EFA目标仓位×0.75" in text
    assert "| 官方腿 |" not in text
    assert "| EMA腿 |" not in text
    assert "| Bias腿 |" not in text
    assert "| LogVol腿 |" not in text


def test_v78_suba_param_tables_list_each_leg_separately():
    module = load_v78_module()
    chunks = []

    module._write_v78_suba_param_tables(chunks.append)
    text = "".join(chunks)

    assert "**全局执行口径**" in text
    assert "**V7.7A原版参数**" in text
    assert "**New A TV1.0参数**" in text
    assert "**Sub-A风控与执行参数**" in text
    assert "V7.8 Sub-A混合方式" not in text


def test_v78_adk_param_tables_list_each_leg_separately():
    module = load_v78_module()
    chunks = []

    module._write_v78_adk_param_tables(chunks.append)
    text = "".join(chunks)

    assert "**全局执行口径**" in text
    assert "**V7.7 ADK参数**" in text
    assert "**New ADK all10 score-hot参数**" in text
    assert "**ADK风控与执行参数**" in text
    assert "V7.8 ADK混合方式" not in text


def test_v78_live_params_source_uses_v78_section_headers_not_legacy_single_leg_labels():
    root = Path(__file__).resolve().parents[1]
    source = (root / "mnt_bot V 7.8 plus.py").read_text(encoding="utf-8")
    live_params = source.split("def _handle_live_params", 1)[1].split("def _handle_signal_history", 1)[0]

    assert "### Sub-A: V7.8双腿综合" in live_params
    assert "V7.8 Sub-A双腿实时状态" in live_params
    assert "V7.8 ADK双腿实时状态" in live_params
    assert "官方腿分窗口动量排名" in live_params
    assert "Sub-A 乖离动量排序" not in live_params
    assert "多配对Top-1状态" not in live_params


def test_v78_live_signal_adk_does_not_duplicate_legacy_single_leg_top3():
    root = Path(__file__).resolve().parents[1]
    source = (root / "mnt_bot V 7.8 plus.py").read_text(encoding="utf-8")
    live_signal = source.split("def _handle_live_signal", 1)[1].split("def _handle_params", 1)[0]

    assert "实时Top-3（若现在收盘，用于判断是否按收盘价执行；策略实际只持有Top-1）" not in live_signal
    assert "若现在收盘的Top-1配对/方向" not in live_signal
    assert "_write_v78_adk_new_leg_then_summary" in live_signal


def test_v78_adk_page_wording_uses_two_leg_and_net_exposure_language():
    module = load_v78_module()
    combined = "\n".join(
        [
            inspect.getsource(module.CombinedStrategyV78._handle_signal),
            inspect.getsource(module.CombinedStrategyV78._handle_live_signal),
            inspect.getsource(module._v78_adk_position_context_labels),
            inspect.getsource(module._write_v78_adk_position_context_note),
        ]
    )

    assert "当前已生效Top-1持仓" not in combined
    assert "当前已生效Top-1:" not in combined
    assert "当前已确认Top-1配对/方向" not in combined
    assert "今日Top-1配对/方向" not in combined
    assert "当前已生效双腿持仓" in combined
    assert "账户级净敞口" in combined


def test_v78_adk_net_exposure_changed_compares_account_level_exposure():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-15", "2026-06-16", "2026-06-17"])
    result = pd.DataFrame(
        {
            "adk_net_asset_exposure": [
                {"SZ50": 0.50, "CYB": -0.50},
                {"SZ50": 0.50, "CYB": -0.50},
                {"SZ50": 0.25, "CYB": -0.50, "ZZ500": 0.25},
            ],
        },
        index=dates,
    )

    assert module._adk_net_exposure_changed(result, 0, 1) is False
    assert module._adk_net_exposure_changed(result, 1, 2) is True
    assert module._adk_net_exposure_changed(result, -2, -1) is True


def test_v78_live_signal_adk_uses_effective_row_and_net_exposure_alert():
    module = load_v78_module()
    source = inspect.getsource(module.CombinedStrategyV78._handle_live_signal)
    adk_block = source.split("### Sub-A-DK: V7.8双子策略", 1)[1].split("_write_volume_warning_panel", 1)[0]

    assert "_dk_intraday3 = cn_unconfirmed and dk_data_is_today and len(cn_dk_result) >= 2" in adk_block
    assert "_dk_effective_idx3 = -2 if _dk_intraday3 else -1" in adk_block
    assert '_dk_effective_holding3 = cn_dk_result["holding"].iloc[_dk_effective_idx3]' in adk_block
    assert "dk_current_name3 = _dk_pos_str(_dk_effective_holding3)" in adk_block
    assert "_adk_net_exposure_changed(cn_dk_result, _dk_effective_idx3, _dk_hypo_idx3)" in adk_block
    assert "_write_v78_adk_net_exposure_table(w, cn_dk_result, _dk_effective_idx3)" in adk_block
    assert '_dk_switch_alert_text(dk_current_name3, _dk_pos_str(hypo_dk)' not in adk_block
    assert "ADK净敞口将变化" in adk_block
    assert "ADK净敞口无变化" in adk_block


def test_v78_signal_adk_uses_net_exposure_not_legacy_single_leg_hypo():
    module = load_v78_module()
    source = inspect.getsource(module.CombinedStrategyV78._handle_signal)
    adk_block = source.split("### Sub-A-DK: V7.8双子策略", 1)[1].split("# ── DK vol-scaling", 1)[0]
    signal_info_block = source.split('signal_info["Sub-A-DK"] = {', 1)[1].split("# ── DK vol-scaling", 1)[0]

    assert "_dk_signal_current_idx" in adk_block
    assert "_dk_signal_target_idx" in adk_block
    assert "_adk_net_exposure_changed(cn_dk_result, _dk_signal_current_idx, _dk_signal_target_idx)" in adk_block
    assert "_dk_switch_alert_text(_dk_effective_name" not in adk_block
    assert "hypo_dk_name" not in adk_block
    assert "_dk_latest_name" not in adk_block
    assert "ADK净敞口已变化" in adk_block
    assert "ADK净敞口无变化" in adk_block
    assert "_adk_net_exposure_signal_text" in signal_info_block
    assert '"V7.8 ADK为双腿component-net；执行优先看账户级净敞口表"' in signal_info_block
    assert '"signal_text": _dk_latest_name' not in signal_info_block
    assert "_dk_pair_display(_dk_latest_pair)" not in signal_info_block


def test_v78_adk_excel_signal_text_uses_net_exposure_summary():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-16", "2026-06-17"])
    result = pd.DataFrame(
        {
            "adk_net_asset_exposure": [
                {"SZ50": 0.50, "CYB": -0.50},
                {"SZ50": 0.25, "CYB": -0.50, "ZZ500": 0.25},
            ],
        },
        index=dates,
    )

    changed = module._adk_net_exposure_signal_text(result, 0, 1)
    unchanged = module._adk_net_exposure_signal_text(result, 0, 0)

    assert changed == "ADK净敞口有变化，按“ADK净敞口”表复核"
    assert unchanged == "ADK净敞口无变化"


def test_v78_adk_display_branches_before_single_leg_volscale_projection():
    module = load_v78_module()
    for method in (module.CombinedStrategyV78._handle_signal, module.CombinedStrategyV78._handle_live_signal):
        source = inspect.getsource(method)
        adk_block = source.split('if "weight" in cn_dk_result.columns', 1)[1].split("_write_v78_adk_new_leg_then_summary", 1)[0]

        assert adk_block.index('"v78_adk_final_exposure" in cn_dk_result.columns') < adk_block.index("_compute_next_vol_scale")


def test_v78_signal_history_routes_adk_to_component_rebalance_extractor():
    module = load_v78_module()
    source = inspect.getsource(module.CombinedStrategyV78._handle_signal_history)
    adk_block = source.split("# ===== Sub-A-DK =====", 1)[1].split("# ===== Sub-B =====", 1)[0]

    assert 'if "v78_adk_final_exposure" in cn_dk_result.columns' in adk_block
    assert "extract_v78_adk_rebalances" in adk_block
    assert adk_block.index("extract_v78_adk_rebalances") < adk_block.index("_split_dk_history_trades")


def test_v78_signal_history_routes_subb_to_effective_rebalance_extractors():
    module = load_v78_module()
    source = inspect.getsource(module.CombinedStrategyV78._handle_signal_history)
    subb_block = source.split("# ===== Sub-B =====", 1)[1].split("def _handle_nav_chart", 1)[0]

    assert "extract_us_rot_rebalances" in subb_block
    assert "extract_subb_volreg_rebalances" in subb_block
    assert "us_rebal = us_period[" not in subb_block


def test_v78_adk_holding_summary_uses_multi_leg_wording():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    result = pd.DataFrame(
        {
            "holding": ["V7.7ADK:SZ50/ZZ1000_1|NewADK:HS300/CYB_1"],
            "top_pair": ["V7.7ADK:SZ50/ZZ1000|NewADK:HS300/CYB"],
            "direction": [1],
            "weight": [0.8],
            "v78_adk_v77_holding": ["SZ50/ZZ1000_1"],
            "v78_adk_new_holding": ["HS300/CYB_1"],
        },
        index=dates,
    )
    chunks = []

    module._write_v78_adk_current_holding_summary(chunks.append, result, -1)
    text = "".join(chunks)

    assert "ADK当前已生效双腿持仓" in text
    assert "综合持仓标识" in text
    assert "综合Top-1" not in text


def test_signal_excel_writes_adk_net_exposure_sheet():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    cn_dk_result = pd.DataFrame(
        {
            "adk_net_asset_exposure": [
                {"SZ50": 0.5, "ZZ1000": -0.25},
            ],
            "cost_basis_note": ["component-net blend; component costs already included"],
        },
        index=dates,
    )

    excel_bytes = module.generate_signal_excel("2026-06-12", {}, [], cn_dk_result=cn_dk_result)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    assert "ADK净敞口" in wb.sheetnames
    ws = wb["ADK净敞口"]
    assert [ws.cell(1, col).value for col in range(1, 5)] == ["指数", "带符号净敞口", "净敞口", "方向"]
    assert ws.cell(2, 1).value == module._dk_leg_name("SZ50")
    assert ws.cell(2, 2).value == 0.5
    assert ws.cell(2, 3).value == 0.5
    assert ws.cell(2, 4).value == "做多"
    assert ws.cell(5, 1).value == "日期口径"
    assert ws.cell(5, 2).value == "当前已生效"
    assert ws.cell(6, 1).value == "净敞口日期"
    assert ws.cell(6, 2).value == "2026-06-12"
    assert ws.cell(8, 1).value == "ADK回测口径"
    assert ws.cell(8, 2).value == "component-net"


def test_signal_excel_adk_net_exposure_sheet_can_use_current_effective_row():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-16", "2026-06-17"])
    cn_dk_result = pd.DataFrame(
        {
            "adk_net_asset_exposure": [
                {"SZ50": 0.50, "CYB": -0.50},
                {"SZ50": 0.25, "CYB": -0.50, "ZZ500": 0.25},
            ],
        },
        index=dates,
    )

    excel_bytes = module.generate_signal_excel(
        "2026-06-17",
        {},
        [],
        cn_dk_result=cn_dk_result,
        adk_net_row_idx=-2,
        adk_net_date_label="当前已生效",
    )
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb["ADK净敞口"]

    assert ws.cell(2, 1).value == module._dk_leg_name("SZ50")
    assert ws.cell(2, 2).value == 0.5
    assert ws.cell(5, 1).value == "日期口径"
    assert ws.cell(5, 2).value == "当前已生效"
    assert ws.cell(6, 1).value == "净敞口日期"
    assert ws.cell(6, 2).value == "2026-06-16"


def test_v78_signal_excel_receives_adk_current_effective_row_idx():
    module = load_v78_module()
    source = inspect.getsource(module.CombinedStrategyV78._handle_signal)
    excel_block = source.split("excel_bytes = generate_signal_excel", 1)[1].split("filename =", 1)[0]

    assert "adk_net_row_idx=_dk_signal_current_idx" in excel_block
    assert 'adk_net_date_label="当前已生效"' in excel_block


def test_v78_adk_legacy_single_leg_switch_alert_and_hypo_key_are_removed():
    module = load_v78_module()
    source = inspect.getsource(module)

    assert "def _dk_switch_alert_text" not in source
    assert '"hypo_dk":' not in source
    assert 'd.get("hypo_dk"' not in source
    assert 'd["hypo_dk"]' not in source


def test_v78_subb_spy_volume_fail_mode_is_parameterized(monkeypatch):
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-11", "2026-06-12"])
    monkeypatch.setattr(
        module,
        "_v78_fetch_spy_volume",
        lambda index: (pd.Series(False, index=index, dtype=bool), "unavailable: test missing"),
    )

    monkeypatch.setattr(module, "V78_SUBB_SPY_VOLUME_FAIL_MODE", "warn_open")
    gate, source = module._v78_spy_volume_gate(dates)
    assert gate.tolist() == [False, False]
    assert source.startswith("unavailable:")

    monkeypatch.setattr(module, "V78_SUBB_SPY_VOLUME_FAIL_MODE", "fail_closed")
    gate, source = module._v78_spy_volume_gate(dates)
    assert gate.tolist() == [True, True]
    assert "fail_closed" in source

    monkeypatch.setattr(module, "V78_SUBB_SPY_VOLUME_FAIL_MODE", "raise")
    try:
        module._v78_spy_volume_gate(dates)
    except RuntimeError as exc:
        assert "SPY volume unavailable" in str(exc)
    else:
        raise AssertionError("raise mode should stop when SPY volume is unavailable")


def test_v78_subb_spy_volume_default_is_fail_closed():
    module = load_v78_module()

    assert module.V78_SUBB_SPY_VOLUME_FAIL_MODE == "fail_closed"


def test_v78_us_yahoo_daily_timestamps_are_parsed_in_utc():
    module = load_v78_module()
    source = inspect.getsource(module._fetch_us_yahoo)

    assert 'pd.Timestamp.fromtimestamp(ts, tz="UTC")' in source
    assert "pd.Timestamp.fromtimestamp(ts).strftime" not in source


def test_v78_spy_volume_fetch_uses_retry_session():
    module = load_v78_module()
    source = inspect.getsource(module._v78_fetch_spy_volume)

    assert "_session.get(" in source
    assert "requests.get(" not in source


def test_v78_subb_btc_start_filter_masks_pre_inception_history():
    module = load_v78_module()
    dates = pd.to_datetime(["2021-12-30", "2022-01-03"])
    close = pd.DataFrame(
        {
            module.US_ROT_BTC_TICKER: [100.0, 110.0],
            "QQQ": [200.0, 210.0],
        },
        index=dates,
    )

    out = module._apply_subb_btc_start_filter(close)

    assert pd.isna(out.loc[dates[0], module.US_ROT_BTC_TICKER])
    assert out.loc[dates[1], module.US_ROT_BTC_TICKER] == 110.0
    assert out.loc[dates[0], "QQQ"] == 200.0


def test_v78_subb_all_rotation_legs_apply_btc_start_filter():
    module = load_v78_module()

    for fn in (
        module.run_us_rotation_mix,
        module.run_subb_v75_ema_base7_rotation,
        module.run_v78_subb_new_line,
    ):
        source = inspect.getsource(fn)
        assert "_apply_subb_btc_start_filter(close_df)" in source


def test_v78_subb_btc_start_filter_is_owned_by_rotation_runners():
    module = load_v78_module()
    source = inspect.getsource(module.CombinedStrategyBase._cached_fetch_data)

    assert "us_rot_close = _apply_subb_btc_start_filter(us_rot_close)" not in source


def test_v78_adk_r2_uses_same_weight_kernel_as_bias_slope():
    module = load_v78_module()
    source = inspect.getsource(module._dk_calc_bias_momentum_r2)

    assert "np.linspace(1.0, 10.0, mom_day)" in source
    assert "np.arange(1, mom_day + 1" not in source


def test_v78_new_adk_primary_does_not_mutate_global_strategy_flags():
    module = load_v78_module()
    source = inspect.getsource(module.run_v78_adk_new_primary)

    assert "globals()[" not in source
    assert "globals().update" not in source
    assert "official_pair_order=_v78_all_adk_pair_order()" in source
    assert "r2_quality_enabled=False" in source


def test_v78_module_has_no_import_time_loop_variable_delete():
    root = Path(__file__).resolve().parents[1]
    source = (root / "mnt_bot V 7.8 plus.py").read_text(encoding="utf-8")

    assert "del _bt_remaining, _n, _c, _dbmf_w, _pre_dbmf_rest, _w" not in source


def test_v78_cn_and_adk_trading_day_basis_are_unified():
    module = load_v78_module()

    assert module.CN_DK_TRADING_DAYS == module.CN_TRADING_DAYS


def test_v78_subb_volume_warning_reports_configured_fail_mode(monkeypatch):
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    result = pd.DataFrame({"return": [0.0]}, index=dates)
    result.attrs["v78_subb_bias"] = pd.DataFrame(
        {"volume_source": ["unavailable: SPY volume missing"]},
        index=dates,
    )
    result.attrs["v78_subb_logvol"] = pd.DataFrame(
        {"volume_source": ["ok"]},
        index=dates,
    )

    monkeypatch.setattr(module, "V78_SUBB_SPY_VOLUME_FAIL_MODE", "warn_open")
    text = module._v78_subb_volume_warning(result)

    assert "SPY volume unavailable" in text
    assert "Bias" in text
    assert "未执行" in text

    monkeypatch.setattr(module, "V78_SUBB_SPY_VOLUME_FAIL_MODE", "fail_closed")
    text = module._v78_subb_volume_warning(result)
    assert "保守降权" in text


def test_signal_excel_uses_plain_no_for_non_signal_rows():
    module = load_v78_module()

    excel_bytes = module.generate_signal_excel(
        "2026-06-12",
        {"Sub-A": {"is_signal": False, "signal_text": "持仓", "note": ""}},
        [],
    )
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb["信号概览"]

    assert ws.cell(2, 2).value == "否"


def _synthetic_subb_close_with_macro_winners(module, periods=460):
    index = pd.bdate_range("2024-01-02", periods=periods)
    close = pd.DataFrame(index=index)
    t = np.arange(periods)
    for i, code in enumerate(module.US_ROT_POOL):
        close[code] = 100.0 * np.exp((0.00005 + i * 0.000001) * t + 0.01 * np.sin(t / (5.0 + i)))
    close["UUP"] = 100.0 * np.exp(0.0018 * t + 0.012 * np.sin(t / 5.0))
    close["DBMF"] = 100.0 * np.exp(0.0020 * t + 0.012 * np.sin(t / 6.0))
    close["KMLM"] = 100.0 * np.exp(0.0019 * t + 0.012 * np.sin(t / 7.0))
    close["DBC"] = 100.0 * np.exp(-0.0007 * t + 0.011 * np.sin(t / 5.5))
    close["TLT"] = 100.0 * np.exp(0.0008 * t + 0.010 * np.sin(t / 6.5))
    close["BIL"] = 100.0 * np.exp(0.00002 * t)
    close["SPY"] = 100.0 * np.exp(0.0004 * t + 0.009 * np.sin(t / 6.0))
    return close


def test_v78_subb_hypothetical_new_leg_weights_read_component_result_attrs():
    module = load_v78_module()
    close = _synthetic_subb_close_with_macro_winners(module)
    dates = close.index
    bias_component = pd.DataFrame(
        {"target_vol_scale": [1.23], "volume_scale_next": [0.75]},
        index=[dates[-1]],
    )
    blended = pd.DataFrame(index=[dates[-1]])
    blended.attrs["v78_subb_bias"] = bias_component

    direct = module._v78_subb_new_line_hypo_weights(close, bias_component, line="bias", row_idx=-1)
    via_attrs = module._v78_subb_new_line_hypo_weights_from_blend(close, blended, line="bias", row_idx=-1)

    assert via_attrs == direct


def test_v78_subb_inflation_note_covers_all_four_legs():
    module = load_v78_module()

    note = module._v78_subb_inflation_participation_note()

    assert "官方腿：通胀开关ON才纳入UUP/DBMF/KMLM" in note
    assert "EMA腿/Bias腿/LogVol腿：始终US_ROT_POOL全池参与排名" in note
    assert "UUP/DBMF/KMLM" in note


def test_v78_subb_default_rule_text_matches_macro_pool_scan_decision():
    module = load_v78_module()

    text = module._v78_subb_default_rule_text()

    assert "官方腿25%" in text
    assert "通胀开关ON才纳入UUP/DBMF/KMLM" in text
    assert "EMA腿25%" in text
    assert "Bias腿25%" in text
    assert "LogVol腿25%" in text
    assert "EMA腿/Bias腿/LogVol腿：始终US_ROT_POOL全池参与排名" in text


def test_v78_subb_inflation_status_text_uses_green_red_markers():
    module = load_v78_module()

    on_text = module._v78_subb_inflation_status_text(True)
    off_text = module._v78_subb_inflation_status_text(False)

    assert "🟢 官方腿宏观池开启" in on_text
    assert "🔴 官方腿宏观池关闭" in off_text
    assert "🟢 EMA/Bias/LogVol全池开启" in on_text
    assert "🟢 EMA/Bias/LogVol全池开启" in off_text


def test_v78_subb_window_lbs_are_safe_for_display_if_global_is_polluted():
    module = load_v78_module()
    original = module.US_ROT_LBS
    try:
        module.US_ROT_LBS = pd.DataFrame({"x": [160, 260, 390]})
        assert module._subb_window_lbs_for_display() == (160, 260, 390)
        assert "160/260/390" in module._v78_subb_default_rule_text()
    finally:
        module.US_ROT_LBS = original


def test_v78_subb_bias_and_logvol_macro_assets_stay_in_full_pool_when_inflation_gate_off():
    module = load_v78_module()
    close = _synthetic_subb_close_with_macro_winners(module)
    module._v78_spy_volume_gate = lambda index: (pd.Series(False, index=index), "test volume")

    assert module._subb_active_ranking_codes(close, -1) == module.US_ROT_BASE_POOL

    bias = module.run_v78_subb_new_line(close, line="bias")
    logvol = module.run_v78_subb_new_line(close, line="logvol")

    for result in (bias, logvol):
        signal_rows = result[result["is_signal"]]
        assert not signal_rows.empty
        last_signal = signal_rows.iloc[-1]
        assert last_signal["target_w_DBMF"] > 0
        assert last_signal["target_w_KMLM"] > 0
        assert last_signal["target_w_UUP"] > 0


def test_v78_adk_rank_rows_are_built_per_leg_not_from_blended_v77_only():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame({"top_pair": ["SZ50/CYB"], "direction": [1]}, index=dates)
    v77.attrs["signals_df"] = pd.DataFrame(
        {"SZ50/CYB": [44.0], "HS300/CYB": [34.0], "SZ50/ZZ1000": [28.0]},
        index=dates,
    )
    new = pd.DataFrame({"top_pair": ["HS300/ZZ500"], "direction": [-1]}, index=dates)
    new.attrs["signals_df"] = pd.DataFrame(
        {"HS300/ZZ500": [99.0], "SZ50/CYB": [44.0], "HS300/CYB": [34.0]},
        index=dates,
    )
    blended = pd.DataFrame({"top_pair": ["SZ50/CYB"], "direction": [1]}, index=dates)
    blended.attrs["v78_adk_v77"] = v77
    blended.attrs["v78_adk_new"] = new

    sections = module._v78_adk_leg_rank_sections(blended, 0, use_shifted=False, top_n=3)

    assert sections[0]["leg"] == "V7.7 ADK"
    assert sections[0]["rows"][0]["pair"] == "SZ50/CYB"
    assert sections[1]["leg"] == module.V78_ADK_NEW_LABEL
    assert sections[1]["rows"][0]["pair"] == "HS300/ZZ500"


def test_v78_adk_display_uses_unified_two_leg_status_and_rank_sections():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-12"])
    v77 = pd.DataFrame({"top_pair": ["SZ50/CYB"], "direction": [1], "weight": [0.6]}, index=dates)
    v77.attrs["signals_df"] = pd.DataFrame(
        {"SZ50/CYB": [44.0], "HS300/CYB": [34.0], "SZ50/ZZ1000": [28.0]},
        index=dates,
    )
    new = pd.DataFrame(
        {
            "top_pair": ["HS300/ZZ500"],
            "direction": [-1],
            "weight": [0.4],
            "v78_score_overheat_scale": [1.0],
            "v78_score_overheat_on": [False],
        },
        index=dates,
    )
    new.attrs["signals_df"] = pd.DataFrame(
        {"HS300/ZZ500": [99.0], "SZ50/CYB": [44.0], "HS300/CYB": [34.0]},
        index=dates,
    )
    blended = pd.DataFrame(
        {
            "v78_adk_v77_holding": ["SZ50/CYB_1"],
            "v78_adk_new_holding": ["HS300/ZZ500_-1"],
            "v78_adk_v77_weight": [0.6],
            "v78_adk_new_weight": [0.4],
            "v78_adk_final_exposure": [0.5],
        },
        index=dates,
    )
    blended.attrs["v78_adk_v77"] = v77
    blended.attrs["v78_adk_new"] = new
    chunks = []

    module._write_v78_adk_new_leg_then_summary(chunks.append, blended, 0, use_shifted=False)
    text = "".join(chunks)

    assert "**V7.8 ADK 混合腿拆分（沿用7.7展示样式）**" in text
    assert "**V7.8 ADK 子策略状态**" in text
    assert "| 腿 | 配对范围 | 分腿Top-1配对/方向 | 排名分数 | 质量过滤 | 风控/过滤 | 腿内敞口 | 组合贡献 |" in text
    assert "| V7.7 ADK | 正式8配对 | 做多 上证50 / 做空 创业板 | `44.00` | R²质控" in text
    assert f"| {module.V78_ADK_NEW_LABEL} | 全10配对 + score-hot | 做多 中证500 / 做空 沪深300 | `99.00`" in text
    assert "**V7.8 ADK 两个子策略Top-3" in text
    assert "V7.7 ADK（正式8配对） 实时Top-3" in text
    assert "New ADK all10 score-hot（全10配对 + score-hot） 实时Top-3" in text
    assert "- 1. **上证50/创业板** | 实时分数 `44.00` | 方向 +1" in text
    assert "- 1. **沪深300/中证500** | 实时分数 `99.00` | 方向 -1" in text
    assert "← 若现在收盘将执行" in text


def test_v78_adk_summary_labels_position_context():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-16"])
    v77 = pd.DataFrame({"top_pair": ["SZ50/CYB"], "direction": [1], "weight": [0.6]}, index=dates)
    v77.attrs["signals_df"] = pd.DataFrame({"SZ50/CYB": [44.0]}, index=dates)
    new = pd.DataFrame(
        {
            "top_pair": ["HS300/ZZ500"],
            "direction": [-1],
            "weight": [0.4],
            "v78_score_overheat_scale": [1.0],
            "v78_score_overheat_on": [False],
        },
        index=dates,
    )
    new.attrs["signals_df"] = pd.DataFrame({"HS300/ZZ500": [99.0]}, index=dates)
    blended = pd.DataFrame(
        {
            "v78_adk_v77_holding": ["SZ50/CYB_1"],
            "v78_adk_new_holding": ["HS300/ZZ500_-1"],
            "v78_adk_v77_weight": [0.6],
            "v78_adk_new_weight": [0.4],
            "v78_adk_final_exposure": [0.5],
        },
        index=dates,
    )
    blended.attrs["v78_adk_v77"] = v77
    blended.attrs["v78_adk_new"] = new
    chunks = []

    module._write_v78_adk_new_leg_then_summary(
        chunks.append,
        blended,
        0,
        use_shifted=False,
        position_context="若现在收盘目标（非当前正式持仓）",
    )
    text = "".join(chunks)

    assert "**ADK持仓口径:** **若现在收盘目标（非当前正式持仓）**" in text
    assert "当前正式持仓以上方“当前已生效双腿持仓”和“账户级净敞口”表为准" in text
    assert "| 腿 | 组合权重 | 若现在收盘双腿配对/方向 | 腿内敞口 | 组合贡献 |" in text
    assert "| 腿 | 配对范围 | 若现在收盘分腿Top-1配对/方向 | 排名分数 | 质量过滤 | 风控/过滤 | 腿内敞口 | 组合贡献 |" in text


def test_v78_adk_current_holding_summary_shows_composite_and_legs():
    module = load_v78_module()
    dates = pd.to_datetime(["2026-06-16"])
    blended = pd.DataFrame(
        {
            "holding": ["SZ50/CYB_1"],
            "top_pair": ["SZ50/CYB"],
            "direction": [1],
            "weight": [0.53],
            "v78_adk_v77_holding": ["SZ50/CYB_1"],
            "v78_adk_new_holding": ["HS300/ZZ500_-1"],
            "v78_adk_v77_weight": [0.53],
            "v78_adk_new_weight": [0.40],
            "v78_adk_final_exposure": [0.465],
        },
        index=dates,
    )
    chunks = []

    module._write_v78_adk_current_holding_summary(chunks.append, blended, 0)
    text = "".join(chunks)

    assert "**ADK当前已生效双腿持仓:** **做多 上证50 / 做空 创业板**" in text
    assert "综合持仓标识: **上证50/创业板** | 方向 +1" in text
    assert "综合Top-1" not in text
    assert "V7.7 ADK原版: 做多 上证50 / 做空 创业板" in text
    assert "New ADK all10 score-hot: 做多 中证500 / 做空 沪深300" in text
    assert "当前已生效总敞口: **0.53x**" in text
